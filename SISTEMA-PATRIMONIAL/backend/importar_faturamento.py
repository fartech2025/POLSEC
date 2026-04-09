#!/usr/bin/env python3
"""
importar_faturamento.py
=======================
Importa dados históricos de faturamento a partir de planilhas Excel (.xlsx).

Formatos suportados
-------------------
1. FATURAMENTO.xlsx  (exportado do sistema Polsec / RES Estoque)
   Colunas esperadas: UNIDADE | MÊS | ANO | CHAMADOS | MÃO DE OBRA | PEÇAS/MATERIAIS | TOTAL
   Ou variações próximas (o script faz match por similaridade).

2. Relatório por unidade única
   Quando a planilha representa uma só unidade, passe --unidade "NOME DA UNIDADE".

Uso
---
    python importar_faturamento.py FATURAMENTO.xlsx --tenant SLUG_DO_TENANT
    python importar_faturamento.py FATURAMENTO.xlsx --tenant polsec --mes 3 --ano 2026
    python importar_faturamento.py FATURAMENTO.xlsx --tenant polsec --unidade "PENITENCIARIA I FRANCO DA ROCHA"
    python importar_faturamento.py --dry-run FATURAMENTO.xlsx --tenant polsec

Dependências (já presentes no venv):
    pip install openpyxl sqlalchemy python-dotenv
"""

import argparse
import os
import re
import sys
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

# ── Path setup ────────────────────────────────────────────────────────────────
sys.path.insert(0, str(Path(__file__).parent))

from dotenv import load_dotenv
load_dotenv()

import openpyxl
from sqlalchemy.orm import Session

from app.database import SessionLocal
from app.models.faturamento import FaturamentoHistorico
from app.models.filial import Filial
from app.models.tenant import Tenant

# ── Mapeamento de cabeçalhos aceitos ─────────────────────────────────────────
# Cada chave é o nome canônico; os valores são padrões regex aceitos.
_HEADER_MAP = {
    "unidade":    r"unidade|unit|filial|localidade|penitenci",
    "mes":        r"^m[eê]s$|^month$",
    "ano":        r"^ano$|^year$",
    "chamados":   r"chamados?|tickets?|os\b|ordens?",
    "mao_obra":   r"m[aã]o.?de.?obra|labor|servi[çc]o",
    "pecas":      r"pe[çc]as?|material|insumo|parts?",
    "total":      r"total|valor.?total|grand.?total",
}

_MESES_NOME = {
    "janeiro": 1, "fevereiro": 2, "março": 3, "marco": 3,
    "abril": 4, "maio": 5, "junho": 6, "julho": 7,
    "agosto": 8, "setembro": 9, "outubro": 10,
    "novembro": 11, "dezembro": 12,
}


def _normalizar(s: str) -> str:
    return re.sub(r"\s+", " ", str(s).strip().lower())


def _mapear_cabecalhos(primeira_linha: tuple) -> dict:
    """Retorna {nome_canonico: indice_coluna} para os cabeçalhos encontrados."""
    mapa = {}
    for idx, celula in enumerate(primeira_linha):
        if celula is None:
            continue
        norm = _normalizar(str(celula))
        for canonico, padrao in _HEADER_MAP.items():
            if re.search(padrao, norm, re.I) and canonico not in mapa:
                mapa[canonico] = idx
    return mapa


def _parse_decimal(valor) -> Decimal:
    if valor is None:
        return Decimal("0")
    s = re.sub(r"[^\d,.\-]", "", str(valor)).replace(",", ".")
    try:
        return Decimal(s) if s else Decimal("0")
    except InvalidOperation:
        return Decimal("0")


def _parse_mes(valor, fallback: Optional[int] = None) -> int:
    if valor is None:
        return fallback or 0
    if isinstance(valor, (int, float)):
        return int(valor)
    norm = _normalizar(str(valor))
    for nome, num in _MESES_NOME.items():
        if nome in norm:
            return num
    try:
        return int(norm)
    except ValueError:
        return fallback or 0


def _buscar_filial(db: Session, tenant_id: str, nome: str) -> Optional[Filial]:
    """Busca filial por nome exato ou substring (case-insensitive)."""
    norm = nome.strip().upper()
    filiais = db.query(Filial).filter(Filial.tenant_id == tenant_id).all()
    # Exata
    for f in filiais:
        if f.nome.strip().upper() == norm:
            return f
    # Substring
    for f in filiais:
        if norm in f.nome.strip().upper() or f.nome.strip().upper() in norm:
            return f
    return None


def importar(
    arquivo: str,
    tenant_slug: str,
    mes_fixo: Optional[int] = None,
    ano_fixo: Optional[int] = None,
    unidade_fixa: Optional[str] = None,
    dry_run: bool = False,
) -> None:
    db: Session = SessionLocal()
    try:
        tenant = db.query(Tenant).filter(Tenant.slug == tenant_slug).first()
        if not tenant:
            print(f"[ERRO] Tenant '{tenant_slug}' não encontrado no banco.")
            return

        wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
        nome_arquivo = Path(arquivo).name
        total_inseridos = 0
        total_ignorados = 0
        total_conflitos = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            linhas = list(ws.iter_rows(values_only=True))
            if not linhas:
                continue

            # Encontra a linha de cabeçalho (primeira com conteúdo)
            header_idx = None
            cabecalhos = {}
            for i, linha in enumerate(linhas):
                cabecalhos = _mapear_cabecalhos(linha)
                if cabecalhos:
                    header_idx = i
                    break

            if header_idx is None:
                print(f"  [ABA '{sheet_name}'] Cabeçalho não reconhecido — pulando.")
                continue

            print(f"\n  [ABA '{sheet_name}'] Cabeçalhos encontrados: {list(cabecalhos.keys())}")

            for linha in linhas[header_idx + 1:]:
                if not any(c is not None for c in linha):
                    continue  # linha vazia

                def col(nome):
                    idx = cabecalhos.get(nome)
                    return linha[idx] if idx is not None else None

                # Unidade
                unidade_raw = str(col("unidade") or unidade_fixa or "").strip()
                if not unidade_raw:
                    continue

                # Período
                mes = _parse_mes(col("mes"), fallback=mes_fixo)
                ano_raw = col("ano")
                ano = int(ano_raw) if ano_raw else (ano_fixo or 0)

                if not (1 <= mes <= 12) or ano < 2020:
                    # tenta extrair do nome da aba: "MAR2026", "03-2026", etc.
                    m = re.search(r"(\d{2})[/\-.]?(\d{4})", sheet_name)
                    if m:
                        mes = mes or int(m.group(1))
                        ano = ano or int(m.group(2))
                    if not (1 <= mes <= 12) or ano < 2020:
                        print(f"    [SKIP] Período inválido (mes={mes}, ano={ano}) na linha: {linha[:4]}")
                        total_ignorados += 1
                        continue

                chamados   = int(col("chamados") or 0)
                mao_obra   = _parse_decimal(col("mao_obra"))
                pecas      = _parse_decimal(col("pecas"))
                total_val  = _parse_decimal(col("total"))
                # Se total não veio, calcula
                if total_val == Decimal("0"):
                    total_val = mao_obra + pecas

                filial = _buscar_filial(db, str(tenant.id), unidade_raw)
                filial_id = filial.id if filial else None
                if not filial:
                    print(f"    [AVISO] Filial '{unidade_raw}' não cadastrada — será importada sem FK.")

                # Verifica conflito
                existe = db.query(FaturamentoHistorico).filter(
                    FaturamentoHistorico.tenant_id == str(tenant.id),
                    FaturamentoHistorico.filial_nome == unidade_raw,
                    FaturamentoHistorico.mes == mes,
                    FaturamentoHistorico.ano == ano,
                    FaturamentoHistorico.origem == "importacao",
                ).first()

                if existe:
                    print(f"    [CONFLITO] {unidade_raw} {mes:02d}/{ano} já existe — pulando.")
                    total_conflitos += 1
                    continue

                registro = FaturamentoHistorico(
                    tenant_id=str(tenant.id),
                    filial_id=filial_id,
                    filial_nome=unidade_raw,
                    mes=mes,
                    ano=ano,
                    chamados_count=chamados,
                    valor_mao_obra=mao_obra,
                    valor_pecas=pecas,
                    valor_total=total_val,
                    origem="importacao",
                    arquivo_origem=nome_arquivo,
                    fechado_em=datetime.utcnow(),
                )
                print(
                    f"    [+] {unidade_raw:40s}  {mes:02d}/{ano}  "
                    f"chamados={chamados:3d}  "
                    f"mao_obra=R${mao_obra:>10.2f}  "
                    f"pecas=R${pecas:>10.2f}  "
                    f"total=R${total_val:>10.2f}"
                )
                if not dry_run:
                    db.add(registro)
                total_inseridos += 1

        if not dry_run:
            db.commit()
            print(f"\n✓ Importação concluída: {total_inseridos} inseridos, "
                  f"{total_conflitos} conflitos pulados, {total_ignorados} ignorados.")
        else:
            print(f"\n[DRY-RUN] {total_inseridos} registros seriam inseridos, "
                  f"{total_conflitos} conflitos, {total_ignorados} ignorados. "
                  "Nenhuma alteração foi feita no banco.")
    finally:
        db.close()


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Importa faturamento histórico de planilhas Excel para o banco POLSEC."
    )
    parser.add_argument("arquivo", help="Caminho para o arquivo .xlsx")
    parser.add_argument("--tenant", required=True, metavar="SLUG",
                        help="Slug do tenant (ex: polsec)")
    parser.add_argument("--mes",  type=int, default=None, metavar="1-12",
                        help="Mês fixo para todas as linhas (quando a planilha não tem coluna MÊS)")
    parser.add_argument("--ano",  type=int, default=None, metavar="YYYY",
                        help="Ano fixo para todas as linhas")
    parser.add_argument("--unidade", default=None, metavar="NOME",
                        help="Nome fixo da unidade (para relatórios de unidade única)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Simula a importação sem gravar no banco")
    args = parser.parse_args()

    if not os.path.isfile(args.arquivo):
        print(f"[ERRO] Arquivo não encontrado: {args.arquivo}")
        sys.exit(1)

    print(f"Importando '{args.arquivo}' → tenant '{args.tenant}'"
          + (" [DRY-RUN]" if args.dry_run else ""))
    importar(
        arquivo=args.arquivo,
        tenant_slug=args.tenant,
        mes_fixo=args.mes,
        ano_fixo=args.ano,
        unidade_fixa=args.unidade,
        dry_run=args.dry_run,
    )


if __name__ == "__main__":
    main()
