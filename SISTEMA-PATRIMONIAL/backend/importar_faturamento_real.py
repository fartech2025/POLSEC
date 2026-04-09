#!/usr/bin/env python3
"""
importar_faturamento_real.py
============================
Importa dados do FATURAMENTO.xlsx (formato real da EMTEL/POLSEC).

Estrutura do arquivo:
  - Uma aba por mês ("Outubro 2022", "MARÇO 2026", etc.)
  - Cada aba tem linhas de unidades prisionais com valor de faturamento mensal
  - Coluna "Valor total de faturamento no mês" detectada dinamicamente por header

Uso:
    python importar_faturamento_real.py FATURAMENTO.xlsx --tenant polsec
    python importar_faturamento_real.py FATURAMENTO.xlsx --tenant polsec --dry-run
    python importar_faturamento_real.py FATURAMENTO.xlsx --tenant polsec --desde 2025-01
    python importar_faturamento_real.py FATURAMENTO.xlsx --tenant polsec --aba "MARÇO 2026"
"""

import argparse
import re
import sys
import warnings
from decimal import Decimal
from pathlib import Path
from typing import Optional

warnings.filterwarnings("ignore")

sys.path.insert(0, str(Path(__file__).parent))

from dotenv import load_dotenv
load_dotenv()

import openpyxl
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.database import SessionLocal
from app.models.faturamento import FaturamentoHistorico
from app.models.filial import Filial
from app.models.tenant import Tenant

# ── Mapeamento de meses em PT-BR ──────────────────────────────────────────────
_MESES = {
    "janeiro": 1, "fevereiro": 2, "março": 3, "marco": 3,
    "abril": 4, "maio": 5, "junho": 6, "julho": 7,
    "agosto": 8, "setembro": 9, "outubro": 10,
    "novembro": 11, "dezembro": 12,
}

# Abas que NÃO contêm dados de faturamento
_ABAS_IGNORAR = {"descrição", "descricao", "glosa", "entregas", "resumo", "indice", "índice"}


def _parse_aba(nome_aba: str) -> Optional[tuple]:
    """Extrai (mes, ano) do nome da aba. Retorna None se não for aba de mês."""
    norm = nome_aba.strip().lower()

    # Ignora abas especiais
    if norm in _ABAS_IGNORAR:
        return None

    # Procura padrão "nome_mes ano" como "Outubro 2022" ou "MARÇO 2026"
    m = re.search(r"([a-zçáéíóúâêîôûã]+)\s+(\d{4})", norm)
    if m:
        mes_nome = m.group(1)
        ano = int(m.group(2))
        mes = _MESES.get(mes_nome)
        if mes and 2020 <= ano <= 2030:
            return (mes, ano)

    return None


def _encontrar_coluna_valor_total(header: tuple) -> int:
    """
    Retorna o índice da coluna 'Valor total de faturamento no mês'.
    Detecta dinamicamente pois a posição muda entre abas.
    """
    for i, cell in enumerate(header):
        if cell is None:
            continue
        norm = str(cell).strip().lower()
        if "valor total" in norm and "faturamento" in norm:
            return i
        if re.match(r"valor\s+total\s+de\s+fatur", norm):
            return i
    # Fallback: procura só "valor total"
    for i, cell in enumerate(header):
        if cell is None:
            continue
        norm = str(cell).strip().lower()
        if "valor total" in norm:
            return i
    return -1


def _is_linha_dados(row: tuple) -> bool:
    """Verifica se a linha contém dados de unidade (não é header, total, ou vazia)."""
    if not any(c is not None for c in row):
        return False
    lote = row[0]
    unidade = row[2] if len(row) > 2 else None
    # Linha de dados tem Lote numérico e Unidade string
    if not isinstance(lote, (int, float)):
        return False
    if not unidade or not isinstance(unidade, str):
        return False
    return True


def _buscar_ou_criar_filial(db: Session, tenant_id: str, nome: str, dry_run: bool) -> Optional[Filial]:
    """Busca filial existente por nome exato. Cria se não existir."""
    nome = nome.strip()

    # Busca exata (case-insensitive)
    filial = db.query(Filial).filter(
        Filial.tenant_id == tenant_id,
        Filial.nome.ilike(nome)
    ).first()

    if filial:
        return filial

    if dry_run:
        print(f"      [DRY-RUN] Criaria filial: '{nome}'")
        return None

    filial = Filial(
        tenant_id=tenant_id,
        nome=nome,
        ativa=True,
    )
    db.add(filial)
    db.flush()  # Obtém o ID sem fazer commit (o commit é feito por aba)
    db.refresh(filial)
    print(f"      [NOVO] Filial criada: '{nome}' (id={filial.id})")
    return filial


def importar(
    arquivo: str,
    tenant_slug: str,
    dry_run: bool = False,
    desde_mes: Optional[int] = None,
    desde_ano: Optional[int] = None,
    aba_filtro: Optional[str] = None,
) -> None:
    db: Session = SessionLocal()
    try:
        tenant = db.query(Tenant).filter(Tenant.slug == tenant_slug).first()
        if not tenant:
            print(f"[ERRO] Tenant '{tenant_slug}' não encontrado.")
            return

        print(f"[OK] Tenant: {tenant.nome} ({tenant.id})")

        # Configura RLS e aumenta timeout para este script de importação
        tenant_id_str = str(tenant.id)
        db.execute(text(f"SET app.current_tenant_id = '{tenant_id_str}'"))
        db.execute(text("SET statement_timeout = '300s'"))
        db.commit()

        wb = openpyxl.load_workbook(arquivo, data_only=True)
        nome_arquivo = Path(arquivo).name

        total_inseridos = 0
        total_atualizados = 0
        total_ignorados = 0
        total_erros = 0
        filiais_criadas = set()

        abas_processar = wb.sheetnames
        if aba_filtro:
            abas_processar = [a for a in abas_processar if a.lower() == aba_filtro.lower()]
            if not abas_processar:
                print(f"[ERRO] Aba '{aba_filtro}' não encontrada. Abas disponíveis: {wb.sheetnames}")
                return

        for nome_aba in abas_processar:
            periodo = _parse_aba(nome_aba)
            if periodo is None:
                continue

            mes, ano = periodo

            # Filtro de período inicial
            if desde_ano and desde_mes:
                if (ano, mes) < (desde_ano, desde_mes):
                    continue

            ws = wb[nome_aba]
            linhas = list(ws.iter_rows(min_row=1, values_only=True))
            if not linhas:
                continue

            # Primeira linha é sempre o header
            header = linhas[0]
            idx_valor_total = _encontrar_coluna_valor_total(header)

            if idx_valor_total == -1:
                print(f"  [ABA '{nome_aba}'] AVISO: coluna 'Valor total' não encontrada — pulando.")
                continue

            print(f"\n  [ABA '{nome_aba}'] mes={mes}, ano={ano}, col_valor_total={idx_valor_total}")

            for row in linhas[1:]:
                if not _is_linha_dados(row):
                    continue

                unidade = row[2].strip()

                # Valor total
                valor_raw = row[idx_valor_total] if idx_valor_total < len(row) else None
                if valor_raw is None or not isinstance(valor_raw, (int, float)):
                    # Sem valor calculado (fórmula não cacheada ou zero)
                    # Usa valor mensal base × dias / 30 como fallback
                    valor_raw = 0

                valor_total = Decimal(str(round(float(valor_raw), 2)))

                # Busca ou cria filial
                filial = _buscar_ou_criar_filial(db, str(tenant.id), unidade, dry_run)
                filial_id = filial.id if filial else None
                filiais_criadas.add(unidade)

                if dry_run:
                    print(f"    [DRY-RUN] {unidade[:50]:<50} {mes:02d}/{ano}  R$ {valor_total:>12,.2f}")
                    total_inseridos += 1
                    continue

                # Verifica se já existe registro de importação para este período/unidade
                existente = db.query(FaturamentoHistorico).filter(
                    FaturamentoHistorico.tenant_id == str(tenant.id),
                    FaturamentoHistorico.filial_nome == unidade,
                    FaturamentoHistorico.mes == mes,
                    FaturamentoHistorico.ano == ano,
                    FaturamentoHistorico.origem == "importacao",
                ).first()

                if existente:
                    # Atualiza se valor mudou
                    if abs(float(existente.valor_total) - float(valor_total)) > 0.01:
                        existente.valor_total = valor_total
                        existente.valor_mao_obra = valor_total  # sem split disponível
                        existente.arquivo_origem = nome_arquivo
                        print(f"    [UPD] {unidade[:50]:<50} {mes:02d}/{ano}  R$ {valor_total:>12,.2f}")
                        total_atualizados += 1
                    else:
                        total_ignorados += 1
                    continue

                # Cria novo registro
                hist = FaturamentoHistorico(
                    tenant_id=str(tenant.id),
                    filial_id=filial_id,
                    filial_nome=unidade,
                    mes=mes,
                    ano=ano,
                    chamados_count=0,
                    valor_mao_obra=valor_total,  # sem split mão de obra/peças na planilha
                    valor_pecas=Decimal("0"),
                    valor_total=valor_total,
                    observacoes=f"Importado de {nome_arquivo} - aba '{nome_aba}'",
                    origem="importacao",
                    arquivo_origem=nome_arquivo,
                    fechado_por_id=None,
                    fechado_em=None,
                )
                db.add(hist)
                print(f"    [INS] {unidade[:50]:<50} {mes:02d}/{ano}  R$ {valor_total:>12,.2f}")
                total_inseridos += 1

            # Commit por aba para evitar timeout
            if not dry_run:
                db.commit()
                # Re-seta RLS após commit (SET SESSION não persiste em transaction pooler)
                db.execute(text(f"SET app.current_tenant_id = '{str(tenant.id)}'"))
                db.execute(text("SET statement_timeout = '300s'"))
                db.commit()
                print(f"    [OK] Commit aba '{nome_aba}' — {total_inseridos + total_atualizados} registros até agora")

        if not dry_run:
            print(f"\n{'='*65}")
            print(f"Importação concluída!")
        else:
            print(f"\n{'='*65}")
            print(f"DRY-RUN — nenhuma alteração no banco.")

        print(f"Inseridos:   {total_inseridos}")
        print(f"Atualizados: {total_atualizados}")
        print(f"Ignorados:   {total_ignorados}")
        print(f"Erros:       {total_erros}")
        print(f"Unidades:    {len(filiais_criadas)}")

    except Exception as e:
        db.rollback()
        print(f"[ERRO FATAL] {e}")
        raise
    finally:
        db.close()


def main():
    parser = argparse.ArgumentParser(description="Importa FATURAMENTO.xlsx para o banco POLSEC")
    parser.add_argument("arquivo", help="Caminho para o arquivo FATURAMENTO.xlsx")
    parser.add_argument("--tenant", required=True, help="Slug do tenant (ex: polsec)")
    parser.add_argument("--dry-run", action="store_true", help="Apenas simula, não grava no banco")
    parser.add_argument("--desde", metavar="AAAA-MM", help="Processa apenas a partir deste mês (ex: 2025-01)")
    parser.add_argument("--aba", metavar="NOME", help="Processa apenas esta aba específica")
    args = parser.parse_args()

    desde_mes = desde_ano = None
    if args.desde:
        m = re.match(r"(\d{4})-(\d{2})", args.desde)
        if m:
            desde_ano, desde_mes = int(m.group(1)), int(m.group(2))
        else:
            print(f"[ERRO] Formato inválido para --desde: use AAAA-MM")
            sys.exit(1)

    importar(
        arquivo=args.arquivo,
        tenant_slug=args.tenant,
        dry_run=args.dry_run,
        desde_mes=desde_mes,
        desde_ano=desde_ano,
        aba_filtro=args.aba,
    )


if __name__ == "__main__":
    main()
